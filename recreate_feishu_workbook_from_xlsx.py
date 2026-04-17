from __future__ import annotations

import argparse
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from feishu_integration import FeishuClient, _column_letter, load_feishu_config
from feishu_workbook import load_reporting_config, save_state
from legacy_feishu_history import extract_cell_text, extract_cell_url, normalize_source_url


URLISH_HEADERS = {
    "来源链接",
    "目标网站",
    "成功链接",
    "当前应发站点URL",
}

URLISH_HEADER_KEYWORDS = ("链接", "URL", "网址", "网站")


def create_feishu_user_client(config_path: str = "config.json") -> FeishuClient:
    config = load_feishu_config(config_path)
    if not config.get("enabled"):
        raise RuntimeError("飞书配置未启用。")
    if not config.get("app_id") or not config.get("app_secret"):
        raise RuntimeError("飞书 app_id / app_secret 未配置。")
    return FeishuClient(
        app_id=config["app_id"],
        app_secret=config["app_secret"],
        spreadsheet_token="",
        sheet_id="",
        auth_mode=config.get("auth_mode", "user"),
        redirect_uri=config.get("redirect_uri", "http://127.0.0.1:8787/callback"),
        user_token_file=config.get("user_token_file", ".feishu_user_token.json"),
        scopes=config.get("scopes"),
        timeout=int(config.get("timeout_seconds", 20) or 20),
        request_retries=int(config.get("request_retries", 3) or 3),
        request_backoff_seconds=float(config.get("request_backoff_seconds", 2) or 2),
    )


def should_treat_as_url(header: str) -> bool:
    text = str(header or "").strip()
    if text in URLISH_HEADERS:
        return True
    return any(keyword.lower() in text.lower() for keyword in URLISH_HEADER_KEYWORDS)


def clean_cell(header: str, value) -> str:
    if value is None:
        return ""
    if should_treat_as_url(header):
        url = extract_cell_url(value) or extract_cell_text(value)
        normalized = normalize_source_url(url)
        return normalized or str(url or "").strip()
    text = extract_cell_text(value)
    return str(text or "").strip()


def drop_duplicate_columns(headers: list[str], rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    if not headers:
        return headers, rows

    keep_indices: list[int] = []
    seen_columns: dict[tuple[str, tuple[str, ...]], int] = {}
    for idx, header in enumerate(headers):
        signature = tuple(row[idx] if idx < len(row) else "" for row in rows)
        key = (str(header or "").strip(), signature)
        if key in seen_columns:
            continue
        seen_columns[key] = idx
        keep_indices.append(idx)

    compact_headers = [headers[idx] for idx in keep_indices]
    compact_rows = [[row[idx] if idx < len(row) else "" for idx in keep_indices] for row in rows]
    return compact_headers, compact_rows


def trim_sheet_data(raw_rows: list[tuple], sheet_name: str) -> tuple[list[str], list[list[str]]]:
    if not raw_rows:
        return [], []

    header_row = list(raw_rows[0] or [])
    data_rows = [list(row or []) for row in raw_rows[1:]]

    max_width = max([len(header_row)] + [len(row) for row in data_rows] + [1])
    headers = [clean_cell(str(cell or "").strip(), cell) for cell in header_row] + [""] * (max_width - len(header_row))
    cleaned_rows: list[list[str]] = []

    non_empty_columns = set()
    for idx, header in enumerate(headers):
        if str(header or "").strip():
            non_empty_columns.add(idx)

    for row in data_rows:
        padded = row + [None] * (max_width - len(row))
        cleaned = [clean_cell(headers[idx] if idx < len(headers) else "", cell) for idx, cell in enumerate(padded)]
        if any(str(cell or "").strip() for cell in cleaned):
            cleaned_rows.append(cleaned)
            for idx, cell in enumerate(cleaned):
                if str(cell or "").strip():
                    non_empty_columns.add(idx)

    keep_indices = sorted(non_empty_columns)
    headers = [headers[idx] for idx in keep_indices]
    cleaned_rows = [[row[idx] for idx in keep_indices] for row in cleaned_rows]
    headers, cleaned_rows = drop_duplicate_columns(headers, cleaned_rows)

    return headers, cleaned_rows


def load_xlsx_sheets(xlsx_path: Path) -> dict[str, tuple[list[str], list[list[str]]]]:
    workbook = load_workbook(xlsx_path, read_only=False, data_only=True)
    result: dict[str, tuple[list[str], list[list[str]]]] = {}
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers, values = trim_sheet_data(rows, sheet_name)
        result[sheet_name] = (headers, values)
    workbook.close()
    return result


def chunked(values: list[list[str]], size: int) -> list[list[list[str]]]:
    return [values[i : i + size] for i in range(0, len(values), size)]


def write_sheet_minimal(
    client: FeishuClient,
    spreadsheet_token: str,
    sheet_id: str,
    headers: list[str],
    rows: list[list[str]],
    chunk_size: int = 500,
) -> None:
    client.attach_spreadsheet(spreadsheet_token, sheet_id)
    total_columns = max(1, len(headers))
    last_col = _column_letter(total_columns)
    all_rows = [headers] + rows

    if not all_rows:
        all_rows = [[""]]

    written = 0
    for batch in chunked(all_rows, chunk_size):
        start_row = written + 1
        end_row = written + len(batch)
        client.write_range(f"{sheet_id}!A{start_row}:{last_col}{end_row}", batch)
        written += len(batch)

    row_count = max(2, len(all_rows))
    client.resize_sheet(sheet_id, row_count=row_count, column_count=total_columns, frozen_row_count=1)


def update_feishu_config(config_path: Path, spreadsheet_token: str, sheet_ids: dict[str, str]) -> None:
    payload = json.loads(config_path.read_text(encoding="utf-8"))
    feishu = payload.setdefault("feishu", {})
    feishu["spreadsheet_token"] = spreadsheet_token
    feishu["sheet_id"] = sheet_ids.get("sources", "")
    config_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="从本地 xlsx 重建飞书工作簿（最小占用模式）")
    parser.add_argument("--xlsx", required=True, help="本地 xlsx 路径")
    parser.add_argument("--config", default="config.json", help="项目配置文件路径")
    parser.add_argument("--title", default="", help="新建飞书工作簿标题")
    parser.add_argument("--no-config-update", action="store_true", help="不回写 config.json 的 feishu token/sheet_id")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    config_path = Path(args.config).resolve()
    reporting_config = load_reporting_config(str(config_path))
    workbook_title = args.title.strip() or reporting_config.get("workbook_title", "外链运营总表")

    if not xlsx_path.exists():
        raise FileNotFoundError(f"未找到 xlsx 文件: {xlsx_path}")

    print(f"📥 读取本地表格: {xlsx_path}")
    sheet_payloads = load_xlsx_sheets(xlsx_path)
    for name, (headers, rows) in sheet_payloads.items():
        print(f"  - {name}: {len(rows)} 行, {len(headers)} 列")

    client = create_feishu_user_client(str(config_path))
    created = client.create_spreadsheet(workbook_title, as_user=True)
    spreadsheet_token = created["spreadsheet_token"]
    spreadsheet_url = created.get("url", "")
    print(f"🆕 已创建飞书工作簿: {spreadsheet_url or spreadsheet_token}")

    title_to_key = {title: key for key, title in reporting_config.get("sheet_titles", {}).items()}
    sheet_ids: dict[str, str] = {}

    first_sheet_id = client.get_sheet_id_by_token(spreadsheet_token, as_user=True)
    source_title = reporting_config["sheet_titles"]["sources"]
    client.rename_sheet(first_sheet_id, source_title, spreadsheet_token=spreadsheet_token, as_user=True)
    sheet_ids["sources"] = first_sheet_id

    for key, title in reporting_config["sheet_titles"].items():
        if key == "sources":
            continue
        sheet_ids[key] = client.ensure_sheet(title, spreadsheet_token=spreadsheet_token, as_user=True)

    for sheet_name, (headers, rows) in sheet_payloads.items():
        if sheet_name not in title_to_key:
            print(f"⚠️ 跳过未知 sheet: {sheet_name}")
            continue
        key = title_to_key[sheet_name]
        print(f"📤 回填 {sheet_name} ...")
        write_sheet_minimal(client, spreadsheet_token, sheet_ids[key], headers, rows)

    state_path = Path(reporting_config["state_file"])
    if state_path.exists():
        backup = state_path.with_name(f"{state_path.stem}.bak.{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        shutil.copy2(state_path, backup)
        print(f"🗂️ 已备份旧 state: {backup}")

    state_payload = {
        "spreadsheet_token": spreadsheet_token,
        "spreadsheet_url": spreadsheet_url,
        "sheet_ids": sheet_ids,
    }
    save_state(state_path, state_payload)
    print(f"💾 已更新 state: {state_path}")

    if not args.no_config_update:
        update_feishu_config(config_path, spreadsheet_token, sheet_ids)
        print(f"🛠️ 已更新配置文件: {config_path}")

    print("✅ 重建完成")
    if spreadsheet_url:
        print(spreadsheet_url)


if __name__ == "__main__":
    main()
